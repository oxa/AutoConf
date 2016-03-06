__author__ = "Guillaume Ladhuie"

import jinja2
from openpyxl import load_workbook

def cast(val, to_type, default=None):
    try:
        return to_type(val)
    except (ValueError,TypeError):
        default=val
        return default

# 0. Specify in templates/config the configuration you want for the devices

# 1. Specify the excel configuration file
l2wb = load_workbook('input/L2.xlsx')
l3wb = load_workbook('input/L3.xlsx')

# 2. Specify column names if renamed
#General
ifname = "interface type"
ifnumber = "interface"
vlanid = "vlan"
description = "description/name"
#L2
allowedvlans="allowed vlans"
po="channel-group"
po_mode="mode"
#L3
ipaddr = "ip address"
subint = "sub int"
autoconf = "auto-conf information"
netmask = "netmask"

valid_interface = ["gigabitethernet","fastethernet","vlan","port-channel","tengigabitethernet","tunnel","loopback","serial"]
devices_list= list(set(l2wb.get_sheet_names())|set(l3wb.get_sheet_names()))

index_l2 = []
index_l3 = []
l2ws = l2wb.active
l3ws = l3wb.active
raw_index_l2= l2ws.rows[0]
raw_index_l3= l3ws.rows[0]

for cell in raw_index_l2:
    index_l2.append(cell.value)
for cell in raw_index_l3:
    index_l3.append(cell.value)

for device in devices_list:
    try:
        l2interfaces=[]
        l2vlans=[]
        l2ws=l2wb.get_sheet_by_name(device)
        for row in l2ws.rows:
            if (row[index_l2.index(ifname)].value==u"vlan"):
                l2vlans.append({
                "vlanid":row[index_l2.index(vlanid)].value,
                "name" :row[index_l2.index(description)].value
                })
            elif row[index_l2.index(ifname)].value in valid_interface :
                 l2interfaces.append({
                "ifname" : row[index_l2.index(ifname)].value,
                 "vlanid":row[index_l2.index(vlanid)].value,
                 "ifnumber":row[index_l2.index(ifnumber)].value,
                 "description" :row[index_l2.index(description)].value,
                 "allowedvlans" :cast(row[index_l2.index(allowedvlans)].value,str),
                 "po":row[index_l2.index(po)].value,
                 "po_mode": row[index_l2.index(po_mode)].value
                     })
    except (KeyError):
            print "No L2 configuration file for device : ",device
    try:
        l3interfaces=[]
        l3ws=l3wb.get_sheet_by_name(device)
        for row in l3ws.rows:
            if row[index_l3.index(ifname)].value in valid_interface :
                l3interfaces.append({
                "ifname" : row[index_l3.index(ifname)].value,
                "vlanid":row[index_l3.index(vlanid)].value,
                "ifnumber":row[index_l3.index(ifnumber)].value,
                "description" :row[index_l3.index(description)].value,
                "ipaddr" : row[index_l3.index(ipaddr)].value,
                "netmask" : row[index_l3.index(netmask)].value,
                "subint" : row[index_l3.index(subint)].value,
                "autoconf": row[index_l3.index(autoconf)].value
                    })
            else:
                print "[Invalid Interface Name] Unknown interface type : ",row[index_l3.index(ifname)].value
    except (KeyError):
         print "No L3 configuration file for device : ",device

    templateLoader = jinja2.FileSystemLoader( searchpath="templates" )
    templateEnv = jinja2.Environment( loader=templateLoader )
    TEMPLATE_FILE = "config"
    template = templateEnv.get_template( TEMPLATE_FILE )
    templateVars={"l2interfaces":l2interfaces,"l2vlans":l2vlans,"l3interfaces":l3interfaces}

    outputText = template.render( templateVars,trim_blocks=True,lstrip_blocks=True )
    with open("conf/"+device+".cfg", "wb") as fh:
        fh.write(outputText)